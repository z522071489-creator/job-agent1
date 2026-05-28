# -*- coding: utf-8 -*-
"""
GitHub Actions 本地执行脚本
负责：
  1. 调用 SCF 云函数（获取采集+AI分析结果）
  2. 生成 Excel 文件
  3. 上传到 GitHub Release
  4. 通过 Server酱 推送微信通知

运行环境：GitHub Actions（依赖 Secrets 中的环境变量）
"""

import json
import os
import sys
import requests
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 从环境变量读取配置
# ─────────────────────────────────────────
SCF_URL = os.environ.get("SCF_URL", "")          # SCF 触发器 HTTP URL
SERVERCHAN_KEY = os.environ.get("SERVERCHAN_KEY", "")  # Server酱 SendKey
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")      # GitHub Actions 自动注入
GITHUB_REPO = os.environ.get("GITHUB_REPOSITORY", "")  # owner/repo


# ─────────────────────────────────────────
# Step 1: 调用 SCF 云函数
# ─────────────────────────────────────────

def call_scf() -> dict:
    """触发 SCF 函数，等待返回结果（最多等 10 分钟）"""
    if not SCF_URL:
        print("❌ SCF_URL 未配置，请在 GitHub Secrets 中设置")
        sys.exit(1)
    
    print(f"📡 调用 SCF 云函数: {SCF_URL[:50]}...")
    try:
        resp = requests.post(
            SCF_URL,
            json={"trigger": "github_actions", "date": str(date.today())},
            timeout=600,  # SCF 最长执行 600 秒
        )
        resp.raise_for_status()
        data = resp.json()
        
        # SCF 返回的 body 可能是字符串
        if isinstance(data.get("body"), str):
            data = json.loads(data["body"])
        
        print(f"✅ SCF 返回：{data.get('stats', {}).get('total_articles', 0)} 篇文章，"
              f"{data.get('stats', {}).get('job_count', 0)} 个有效岗位")
        return data
    except requests.exceptions.Timeout:
        print("❌ SCF 调用超时（>600s），请检查函数设置")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 调用 SCF 失败: {e}")
        sys.exit(1)


# ─────────────────────────────────────────
# Step 2: 生成 Excel
# ─────────────────────────────────────────

COLUMNS = [
    ("匹配度", 8),
    ("投递优先级", 12),
    ("岗位名称", 22),
    ("招聘单位", 20),
    ("工作城市", 10),
    ("学历要求", 10),
    ("招聘类型", 12),
    ("发布日期", 12),
    ("投递链接", 35),
    ("薪资信息", 14),
    ("截止日期", 12),
    ("岗位摘要", 40),
    ("原文链接", 35),
    ("来源", 18),
]

PRIORITY_LABEL = {
    True: "✅ 优先投递",
    False: "⚠️ 待复核",
}


def score_to_priority(score: int) -> str:
    return "✅ 优先投递" if score >= 80 else "⚠️ 待复核"


def make_excel(jobs: list[dict], output_path: str) -> str:
    """生成格式化 Excel，返回文件路径"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"招聘岗位_{date.today()}"

    # ── 样式定义 ──
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B5797")
    priority_fill = PatternFill("solid", fgColor="E2EFDA")   # 绿底：优先
    review_fill = PatternFill("solid", fgColor="FFF2CC")     # 黄底：待复核
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── 标题行 ──
    headers = [col[0] for col in COLUMNS]
    widths = [col[1] for col in COLUMNS]
    ws.append(headers)
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # ── 数据行 ──
    for row_idx, job in enumerate(jobs, 2):
        score = job.get("score", 0)
        priority = score_to_priority(score)
        row_fill = priority_fill if score >= 80 else review_fill

        row_data = [
            score,
            priority,
            job.get("job_title", ""),
            job.get("company", ""),
            job.get("city", ""),
            job.get("education", ""),
            job.get("job_type", ""),
            job.get("pub_date", ""),
            job.get("article_url", ""),
            job.get("salary") or "未注明",
            job.get("deadline") or "未注明",
            job.get("summary", ""),
            job.get("article_url", ""),
            job.get("source", ""),
        ]
        ws.append(row_data)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx in (1, 2, 5, 6, 7, 8, 10, 11):
                cell.alignment = center
            else:
                cell.alignment = left
            # 投递链接设为超链接
            if col_idx == 9 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[row_idx].height = 36

    # ── 冻结首行 ──
    ws.freeze_panes = "A2"

    # ── 自动筛选 ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(jobs)+1}"

    # ── 摘要 sheet ──
    ws2 = wb.create_sheet("运行摘要")
    ws2.append(["字段", "值"])
    ws2.append(["生成日期", str(date.today())])
    ws2.append(["有效岗位数", len(jobs)])
    ws2.append(["优先投递数", sum(1 for j in jobs if j.get("score", 0) >= 80)])
    ws2.append(["待复核数", sum(1 for j in jobs if 50 <= j.get("score", 0) < 80)])

    wb.save(output_path)
    print(f"✅ Excel 已生成：{output_path}（{len(jobs)} 行）")
    return output_path


# ─────────────────────────────────────────
# Step 3: 上传到 GitHub Release
# ─────────────────────────────────────────

def upload_to_github_release(file_path: str, today: str) -> str:
    """创建/更新每日 Release，上传 Excel，返回下载 URL"""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        print("⚠️ GitHub 环境变量未设置，跳过上传")
        return ""

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }
    tag = f"daily-{today}"
    release_name = f"招聘岗位 {today}"

    api_base = f"https://api.github.com/repos/{GITHUB_REPO}"

    # 查找或创建 Release
    release_id = None
    upload_url = None
    try:
        r = requests.get(f"{api_base}/releases/tags/{tag}", headers=headers, timeout=15)
        if r.status_code == 200:
            release_id = r.json()["id"]
            upload_url = r.json()["upload_url"].split("{")[0]
            # 删除旧资产
            assets = requests.get(f"{api_base}/releases/{release_id}/assets", headers=headers).json()
            for asset in assets:
                requests.delete(f"{api_base}/releases/assets/{asset['id']}", headers=headers)
            print(f"♻️ 复用已有 Release: {tag}")
        else:
            # 新建 Release
            payload = {
                "tag_name": tag,
                "name": release_name,
                "body": f"自动生成的每日招聘汇总，日期：{today}",
                "prerelease": True,
            }
            r = requests.post(f"{api_base}/releases", headers=headers, json=payload, timeout=15)
            r.raise_for_status()
            release_id = r.json()["id"]
            upload_url = r.json()["upload_url"].split("{")[0]
            print(f"✅ 新建 Release: {tag}")
    except Exception as e:
        print(f"❌ 操作 Release 失败: {e}")
        return ""

    # 上传文件
    try:
        filename = Path(file_path).name
        with open(file_path, "rb") as f:
            file_data = f.read()
        upload_headers = {
            **headers,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        r = requests.post(
            f"{upload_url}?name={filename}",
            headers=upload_headers,
            data=file_data,
            timeout=60,
        )
        r.raise_for_status()
        download_url = r.json()["browser_download_url"]
        print(f"✅ 文件已上传：{download_url}")
        return download_url
    except Exception as e:
        print(f"❌ 上传文件失败: {e}")
        return ""


# ─────────────────────────────────────────
# Step 4: Server酱 推送微信
# ─────────────────────────────────────────

def push_wechat(stats: dict, jobs: list[dict], excel_url: str, today: str):
    """通过 Server酱 推送微信"""
    if not SERVERCHAN_KEY:
        print("⚠️ SERVERCHAN_KEY 未设置，跳过推送")
        return

    priority_jobs = [j for j in jobs if j.get("score", 0) >= 80]
    review_jobs = [j for j in jobs if 50 <= j.get("score", 0) < 80]

    # 构建推送标题
    title = f"📋 {today} 招聘日报 | {len(priority_jobs)} 个优先岗位"

    # 构建推送正文（Markdown）
    lines = [
        f"## 今日采集汇总 ({today})",
        f"",
        f"| 指标 | 数值 |",
        f"|------|------|",
        f"| 采集文章数 | {stats.get('total_articles', 0)} 篇 |",
        f"| 有效岗位数 | {stats.get('job_count', 0)} 个 |",
        f"| ✅ 优先投递 | **{len(priority_jobs)} 个** |",
        f"| ⚠️ 待复核 | {len(review_jobs)} 个 |",
        f"",
    ]

    if excel_url:
        lines += [f"📥 **[点击下载Excel]({excel_url})**", ""]

    if priority_jobs:
        lines += ["## ✅ 优先投递岗位", ""]
        for j in priority_jobs[:10]:  # 最多显示10个
            salary = j.get("salary") or "薪资面议"
            deadline = j.get("deadline") or "未注明截止"
            lines.append(
                f"**{j.get('job_title', '')}** — {j.get('company', '')} | "
                f"{j.get('city', '')} | {salary} | 截止:{deadline} | "
                f"🎯{j.get('score', 0)}分"
            )
        if len(priority_jobs) > 10:
            lines.append(f"... 还有 {len(priority_jobs) - 10} 个，详见Excel")
        lines.append("")

    if stats.get("errors"):
        lines += ["## ⚠️ 采集异常", ""]
        for err in stats["errors"][:5]:
            lines.append(f"- {err}")

    desp = "\n".join(lines)

    # 调用 Server酱 API
    try:
        # Server酱 Turbo 版
        if SERVERCHAN_KEY.startswith("SCT"):
            url = f"https://sctapi.ftqq.com/{SERVERCHAN_KEY}.send"
        else:
            url = f"https://sc.ftqq.com/{SERVERCHAN_KEY}.send"

        resp = requests.post(url, data={"title": title, "desp": desp}, timeout=15)
        result = resp.json()
        if result.get("errno") == 0 or result.get("code") == 0:
            print("✅ 微信推送成功")
        else:
            print(f"⚠️ 微信推送返回异常: {result}")
    except Exception as e:
        print(f"❌ 微信推送失败: {e}")


# ─────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────

def main():
    today = str(date.today())
    print(f"\n{'='*50}")
    print(f"  招聘信息采集系统  {today}")
    print(f"{'='*50}\n")

    # Step 1: 调用 SCF
    result = call_scf()
    jobs = result.get("jobs", [])
    stats = result.get("stats", {})

    if not jobs:
        print("⚠️ 今日无符合条件的岗位")
        if SERVERCHAN_KEY:
            push_wechat(stats, [], "", today)
        sys.exit(0)

    # Step 2: 生成 Excel
    excel_path = f"/tmp/招聘汇总_{today}.xlsx"
    make_excel(jobs, excel_path)

    # Step 3: 上传 GitHub Release
    excel_url = upload_to_github_release(excel_path, today)

    # Step 4: 推送微信
    push_wechat(stats, jobs, excel_url, today)

    print(f"\n✅ 全流程完成！共处理 {len(jobs)} 个岗位")


if __name__ == "__main__":
    main()
