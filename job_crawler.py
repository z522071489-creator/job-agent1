#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
春招岗位自动采集智能体
每天 09:00 自动运行，采集公众号招聘文章，AI筛选打分，生成Excel，推送微信
"""

import os
import re
import json
import time
import requests
from datetime import datetime, date, timedelta
from io import BytesIO
from urllib.parse import quote, urlparse, parse_qs
from pathlib import Path

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image

# ============================================================
# 配置区 —— 通过 GitHub Secrets 注入
# ============================================================
KIMI_API_KEY   = os.environ.get("KIMI_API_KEY", "")
SERVERCHAN_KEY = os.environ.get("SERVERCHAN_KEY", "")

ACCOUNTS = [
    "国企指南", "三晋国企", "山西国企直聘", "国聘行动", "优聘计划",
    "晋招聘", "三晋招聘", "山西招聘汇总", "晋师招聘",
    "山西热门招聘", "太原人才大市场公司招聘找工作",
]

TITLE_INCLUDE_KW = ["招聘", "春招", "校招", "岗位", "录用", "招录", "公告", "招募", "用人"]
TITLE_EXCLUDE_KW = ["实习", "兼职", "劳务", "外包"]
SCORE_THRESHOLD  = 50

# ============================================================
# 个人求职条件 —— 通过 GitHub Secrets 注入
# ============================================================
# 学历：本科及以下（排除：硕士、博士、研究生）
MAX_EDUCATION = "本科"
EXCLUDE_EDUCATION = ["硕士", "研究生", "博士", "硕士研究生", "博士研究生"]

# 性别：男性（排除：限女性的岗位）
PREFERRED_GENDER = "不限"
EXCLUDE_GENDER = ["女性", "女"]

# 专业：计算机类相关专业（排除：非计算机类）
PREFERRED_MAJORS = ["计算机", "软件工程", "网络工程", "信息技术", "IT", "编程", "开发"]
EXCLUDE_MAJORS = []  # 暂不设置强制排除，但会在打分时降分

# 年龄：25岁及以上（排除：小于25岁的岗位）
MIN_AGE = 25
EXCLUDE_YOUNG = ["应届", "25岁以下", "24岁"]

SOGOU_BASE = "https://weixin.sogou.com"
STATE_FILE = "/tmp/job_crawler_state.json"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Referer": "https://weixin.sogou.com/",
}


def fix_url(url: str) -> str:
    """修复搜狗相对路径链接"""
    if not url:
        return ""
    if url.startswith("/link?") or url.startswith("/weixin?"):
        return SOGOU_BASE + url
    return url


def get_time_range() -> tuple[datetime, datetime]:
    """
    获取文章时间范围
    第一次运行：获取7天前至今的文章
    后续运行：获取前一天的文章
    """
    if Path(STATE_FILE).exists():
        try:
            with open(STATE_FILE, "r") as f:
                state = json.load(f)
                last_run = datetime.fromisoformat(state.get("last_run", ""))
                return last_run, datetime.now()
        except:
            pass
    
    return datetime.now() - timedelta(days=7), datetime.now()


def save_state():
    """保存运行状态"""
    with open(STATE_FILE, "w") as f:
        json.dump({"last_run": datetime.now().isoformat()}, f)


def parse_publish_date(date_str: str) -> datetime:
    """解析公众号发布日期字符串"""
    if not date_str:
        return datetime.now()
    
    try:
        if "天前" in date_str:
            days = int(re.search(r"(\d+)天前", date_str).group(1))
            return datetime.now() - timedelta(days=days)
        elif "小时前" in date_str:
            hours = int(re.search(r"(\d+)小时前", date_str).group(1))
            return datetime.now() - timedelta(hours=hours)
        elif "分钟前" in date_str:
            minutes = int(re.search(r"(\d+)分钟前", date_str).group(1))
            return datetime.now() - timedelta(minutes=minutes)
        elif "昨天" in date_str:
            return datetime.now() - timedelta(days=1)
        return datetime.fromisoformat(date_str.split()[0])
    except:
        return datetime.now()


def is_article_valid_time(pub_date_str: str, start_time: datetime, end_time: datetime) -> bool:
    """检查文章发布时间是否在范围内"""
    pub_time = parse_publish_date(pub_date_str)
    return start_time <= pub_time <= end_time


def validate_url_accessible(url: str, timeout: int = 8) -> bool:
    """验证URL是否可访问且有内容"""
    if not url or not url.startswith("http"):
        return False
    try:
        resp = requests.head(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        return resp.status_code == 200
    except:
        try:
            resp = requests.get(url, headers=HEADERS, timeout=timeout)
            return resp.status_code == 200
        except:
            return False


# ============================================================
# 1. 搜狗搜索获取文章列表
# ============================================================
def fetch_articles_sogou(account: str, start_time: datetime, end_time: datetime) -> list[dict]:
    articles = []
    seen_urls = set()
    queries = [f"{account} 招聘", f"{account} 岗位", f"{account} 公告"]

    for q in queries:
        url = f"{SOGOU_BASE}/weixin?type=2&query={quote(q)}&ie=utf8"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=12)
            soup = BeautifulSoup(resp.text, "lxml")
            items = soup.select(".news-list li")

            for item in items:
                title_el = item.select_one("h3 a")
                date_el  = item.select_one(".s2")
                src_el   = item.select_one(".account")
                if not title_el:
                    continue
                
                source = src_el.get_text(strip=True) if src_el else ""
                if account not in source and source:
                    continue
                
                pub_date_str = date_el.get_text(strip=True) if date_el else ""
                if not is_article_valid_time(pub_date_str, start_time, end_time):
                    continue
                
                art_url = fix_url(title_el.get("href", ""))
                if not art_url or art_url in seen_urls:
                    continue
                
                if not validate_url_accessible(art_url):
                    print(f"    ⚠️ URL无法访问: {art_url[:50]}...")
                    continue
                
                seen_urls.add(art_url)
                articles.append({
                    "title":    title_el.get_text(strip=True),
                    "url":      art_url,
                    "pub_date": pub_date_str,
                    "source":   account,
                })
                if len(articles) >= 15:
                    break
            time.sleep(1.5)
        except Exception as e:
            print(f"  [搜狗] {account} 失败: {e}")
        if len(articles) >= 15:
            break
    return articles[:15]


# ============================================================
# 2. 标题关键词过滤
# ============================================================
def filter_by_title(articles: list[dict]) -> list[dict]:
    result = []
    for art in articles:
        t = art["title"]
        if any(kw in t for kw in TITLE_INCLUDE_KW):
            if not any(kw in t for kw in TITLE_EXCLUDE_KW):
                result.append(art)
    return result


# ============================================================
# 3. 抓取文章正文
# ============================================================
def fetch_article_detail(url: str) -> tuple[str, list[str]]:
    """抓取文章正文和图片"""
    if not url.startswith("http"):
        return "", []
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            return "", []
        
        soup = BeautifulSoup(resp.text, "lxml")
        body = soup.select_one("#js_content, .rich_media_content")
        if not body:
            text = soup.get_text()[:5000]
            if len(text) < 100:
                return "", []
            return text, []
        
        text = body.get_text(separator="\n", strip=True)[:8000]
        if len(text) < 100:
            return "", []
        
        imgs = []
        for img in body.select("img"):
            src = img.get("data-src") or img.get("src", "")
            if src and src.startswith("http") and "mmbiz" in src:
                imgs.append(src)
        
        print(f"    ✓ 正文长度: {len(text)} 字，图片: {len(imgs)} 张")
        return text, imgs[:12]
    except Exception as e:
        print(f"    ✗ 抓取正文失败: {e}")
        return "", []


# ============================================================
# 4. 二维码识别和链接提取
# ============================================================
def decode_qr_from_images(img_urls: list[str]) -> list[str]:
    """识别二维码并提取链接"""
    links = []
    try:
        from pyzbar.pyzbar import decode as pyzbar_decode
    except ImportError:
        print("    ⚠️ pyzbar未安装，无法识别二维码")
        return links
    
    for img_url in img_urls:
        try:
            resp = requests.get(img_url, headers=HEADERS, timeout=8)
            img = Image.open(BytesIO(resp.content)).convert("RGB")
            for code in pyzbar_decode(img):
                data = code.data.decode("utf-8", errors="ignore").strip()
                if data.startswith("http"):
                    links.append(data)
                    print(f"    🔳 识别二维码链接: {data[:60]}...")
        except Exception as e:
            pass
    
    return links


def extract_apply_link_from_text(text: str) -> list[str]:
    """从正文中提取投递链接（邮箱、表单链接等）"""
    links = []
    
    # 提取邮箱地址
    emails = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
    for email in emails:
        links.append(email)
        print(f"    📧 提取邮箱: {email}")
    
    # 提取 http/https 链接
    urls = re.findall(r'https?://[^\s）\)]+', text)
    for url in urls:
        url = url.rstrip('.,;:!?）\)')
        if url not in links:
            links.append(url)
            print(f"    🔗 提取链接: {url[:60]}...")
    
    return links


def fetch_apply_page_content(apply_link: str) -> str:
    """获取投递页面内容（用于二次筛选）"""
    if not apply_link.startswith("http"):
        return ""
    
    try:
        resp = requests.get(apply_link, headers=HEADERS, timeout=15)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "lxml")
            text = soup.get_text()[:3000]
            print(f"    📄 投递页面长度: {len(text)} 字")
            return text
    except Exception as e:
        pass
    
    return ""


# ============================================================
# 5. Kimi API 语义筛选 + 打分
# ============================================================
SYSTEM_PROMPT = """你是一个严谨的招聘信息分析助手。你的任务是从招聘文章中提取真实存在的岗位信息，并进行精准评分。

## 基础筛选条件（必须满足，否则excluded=true）
1. **招聘类型必须是社会招聘或校招**
   - 排除：仅限应届生、实习生、志愿者、劳务派遣、外包、临时工
    
2. **学历要求必须是本科及以下**
   - 排除：仅限硕士、博士、研究生的岗位（除非同时招本科）
    
3. **工作性质必须是正式编制**
   - 排除：实习、兼职、劳务、临时、外包

4. **性别要求必须不限制或要求男性**
   - 排除：限女性的岗位

5. **年龄要求必须不排除25岁以上**
   - 排除：明确说明需要25岁以下、应届毕业生等的岗位

## 专业匹配打分规则（总分100）

### 工作地点（基础）
- 工作地点在太原：+40分
- 工作地点在吕梁/山西其他地市：+30分
- 工作地点在其他省份：+5分

### 企业性质（重要）
- 国企/央企/事业单位/大型上市公司：+20分
- 其他企业：+10分

### 岗位匹配（核心）
- 岗位为计算机/IT/开发/软件相关（高度匹配）：+25分
- 岗位为技术类但非计算机专业（中度匹配）：+10分
- 岗位为非技术类（低匹配）：0分

### 其他加分
- 有具体岗位名称（不是"若干"）：+5分
- 有具体薪资范围：+5分
- 有具体招聘截止日期：+5分

## 关键要求
1. **岗位名称必须来自原文**，不要编造或推测
2. **城市信息必须来自原文**，不确定填"待确认"
3. **薪资、学历、截止日期必须来自原文**，没有就填空字符串
4. **同一篇文章可提取多个岗位**，但每个岗位必须真实存在于文章中
5. **如果文章太短或内容无法判断**，返回空的jobs列表

## 输出格式（仅输出纯JSON）
```json
{
  "jobs": [
    {
      "job_name": "岗位名称（必须来自原文）",
      "employer": "招聘单位全称（来自原文）",
      "city": "工作城市（来自原文，不确定填'待确认'）",
      "recruitment_type": "社会招聘/校园招聘/混合招聘",
      "education_req": "学历要求（来自原文）",
      "gender_req": "性别要求（来自原文，不限则填'不限'）",
      "age_req": "年龄要求（来自原文，没有填空字符串）",
      "major_req": "专业要求（来自原文，没有填空字符串）",
      "score": 75,
      "salary_info": "薪资（来自原文，如'15-25k'，没有填空字符串）",
      "deadline": "截止日期（来自原文，没有填空字符串）",
      "summary": "80字以内的岗位职责摘要（必须基于原文）",
      "apply_link": "投递链接（来自原文，没有填空字符串）",
      "excluded": false,
      "exclude_reason": ""
    }
  ]
}
```"""


def analyze_with_kimi(article: dict, apply_links: list[str], apply_page_contents: list[str]) -> list[dict]:
    """AI分析提取岗位信息"""
    if not KIMI_API_KEY:
        print("  ⚠️  KIMI_API_KEY 未配置")
        return []

    content = article.get("content", "")
    if len(content) < 100:
        print("  ⚠️  正文太短（<100字），跳过AI分析")
        return []

    # 构建投递链接和页面内容的信息
    apply_info = ""
    if apply_links:
        apply_info += "\n\n【投递方式】\n" + "\n".join(apply_links)
    if apply_page_contents:
        apply_info += "\n\n【投递页面要求摘要】\n" + "\n".join(apply_page_contents[:3])

    user_msg = (
        f"【文章基本信息】\n"
        f"来源公众号：{article['source']}\n"
        f"文章标题：{article['title']}\n"
        f"发布日期：{article['pub_date']}\n"
        f"原文链接：{article['url']}\n\n"
        f"【文章正文】\n{content}"
        f"{apply_info}\n\n"
        f"请严格按照系统提示分析此招聘文章，仅提取真实存在的岗位。"
        f"特别关注学历、性别、年龄等限制条件。"
        f"如果无法确认岗位信息来自原文，返回空jobs列表。"
    )

    try:
        resp = requests.post(
            "https://api.moonshot.cn/v1/chat/completions",
            headers={"Authorization": f"Bearer {KIMI_API_KEY}", "Content-Type": "application/json"},
            json={
                "model": "moonshot-v1-32k",
                "messages": [
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": user_msg},
                ],
                "temperature": 0.1,
                "max_tokens":  2048,
            },
            timeout=40,
        )
        raw = resp.json()["choices"][0]["message"]["content"]
        
        m = re.search(r"\{[\s\S]*\}", raw)
        if not m:
            print(f"    ⚠️ Kimi返回无法解析")
            return []
        
        data = json.loads(m.group())
        jobs = data.get("jobs", [])
        
        # 验证和处理提取的岗位
        valid_jobs = []
        for job in jobs:
            if job.get("excluded"):
                print(f"    ❌ 排除: {job.get('job_name', '未命名')} - {job.get('exclude_reason', '')}")
                continue
            
            if not job.get("job_name"):
                print(f"    ❌ 岗位名称缺失，已过滤")
                continue
            
            job["source"]        = article["source"]
            job["article_url"]   = fix_url(article["url"])
            job["pub_date"]      = article["pub_date"]
            job["article_title"] = article["title"]
            job["apply_links"]   = apply_links
            valid_jobs.append(job)
        
        print(f"   ✓ 提取: {len(jobs)} 个岗位，有效: {len(valid_jobs)} 个")
        return valid_jobs
        
    except Exception as e:
        print(f"    ✗ Kimi调用失败: {e}")
        return []


def secondary_filter_jobs(jobs: list[dict]) -> list[dict]:
    """
    二次筛选：根据个人求职条件筛选岗位
    排除条件：
    - 学历要求大于本科
    - 限女性岗位
    - 明确要求小于25岁（应届）
    """
    filtered_jobs = []
    excluded_count = 0
    
    for job in jobs:
        exclude_reason = ""
        
        # 1. 检查学历要求
        education = job.get("education_req", "").lower()
        for edu_kw in EXCLUDE_EDUCATION:
            if edu_kw.lower() in education:
                exclude_reason = f"学历要求为 {education}，超过本科"
                excluded_count += 1
                print(f"    ⚠️ 排除: {job.get('job_name')} - {exclude_reason}")
                break
        
        # 2. 检查性别要求
        if not exclude_reason:
            gender = job.get("gender_req", "").lower()
            for gender_kw in EXCLUDE_GENDER:
                if gender_kw.lower() in gender:
                    exclude_reason = f"性别要求为 {gender}"
                    excluded_count += 1
                    print(f"    ⚠️ 排除: {job.get('job_name')} - {exclude_reason}")
                    break
        
        # 3. 检查年龄要求
        if not exclude_reason:
            age = job.get("age_req", "").lower()
            for age_kw in EXCLUDE_YOUNG:
                if age_kw.lower() in age:
                    exclude_reason = f"年龄要求不符合: {age}"
                    excluded_count += 1
                    print(f"    ⚠️ 排除: {job.get('job_name')} - {exclude_reason}")
                    break
        
        if exclude_reason:
            job["excluded"] = True
            job["exclude_reason"] = exclude_reason
        else:
            filtered_jobs.append(job)
    
    if excluded_count > 0:
        print(f"   📌 共排除 {excluded_count} 个岗位")
    
    return filtered_jobs


# ============================================================
# 6. 生成 Excel
# ============================================================
def generate_excel(all_jobs: list[dict]) -> str:
    valid = [j for j in all_jobs if not j.get("excluded") and j.get("score", 0) >= SCORE_THRESHOLD]
    valid.sort(key=lambda x: x.get("score", 0), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"岗位清单_{date.today().strftime('%m%d')}"

    headers = [
        "匹配度", "投递优先级", "岗位名称", "招聘单位", "工作城市",
        "学历要求", "性别要求", "年龄要求", "专业要求",
        "招聘类型", "发布日期", "投递链接",
        "薪资信息", "岗位摘要", "原文链接", "来源公众号",
    ]
    h_fill = PatternFill("solid", fgColor="2E4057")
    h_font = Font(color="FFFFFF", bold=True, size=11, name="微软雅黑")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = h_fill
        c.font = h_font
        c.alignment = center
    ws.row_dimensions[1].height = 32

    fills = {
        "✅ 优先投递": PatternFill("solid", fgColor="D4EDDA"),
        "⚠️ 待复核":  PatternFill("solid", fgColor="FFF3CD"),
    }

    for ri, job in enumerate(valid, 2):
        score    = job.get("score", 0)
        priority = "✅ 优先投递" if score >= 80 else "⚠️ 待复核"
        apply_link = job.get("apply_link", "")
        if not apply_link and job.get("apply_links"):
            apply_link = job["apply_links"][0]
        if not apply_link:
            apply_link = job.get("article_url", "")
        apply_link = fix_url(apply_link)

        row = [
            score, priority,
            job.get("job_name", ""),
            job.get("employer", ""),
            job.get("city", ""),
            job.get("education_req", ""),
            job.get("gender_req", ""),
            job.get("age_req", ""),
            job.get("major_req", ""),
            job.get("recruitment_type", ""),
            job.get("pub_date", ""),
            apply_link,
            job.get("salary_info", ""),
            job.get("summary", ""),
            fix_url(job.get("article_url", "")),
            job.get("source", ""),
        ]
        rf = fills.get(priority)
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if rf:
                c.fill = rf
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if ci == 1:
                c.font = Font(bold=True, size=12)
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 45

    widths = [8, 13, 20, 18, 10, 10, 10, 10, 12, 10, 10, 30, 12, 35, 35, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    path = f"/tmp/招聘岗位_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(path)
    return path


# ============================================================
# 7. 上传文件获取下载链接
# ============================================================
def upload_excel(path: str) -> str:
    try:
        with open(path, "rb") as f:
            resp = requests.post(
                "https://file.io",
                files={"file": (os.path.basename(path), f)},
                data={"expires": "7d"},
                timeout=20,
            )
        data = resp.json()
        if data.get("success"):
            return data.get("link", "")
    except Exception as e:
        print(f"  文件上传失败: {e}")
    return ""


# ============================================================
# 8. Server酱推送微信
# ============================================================
def push_wechat(title: str, body: str):
    if not SERVERCHAN_KEY:
        print("  ⚠️  SERVERCHAN_KEY未配置")
        return
    try:
        resp = requests.post(
            f"https://sctapi.ftqq.com/{SERVERCHAN_KEY}.send",
            data={"title": title, "desp": body},
            timeout=12,
        )
        r = resp.json()
        print("  ✅ 微信推送成功" if r.get("code") == 0 else f"  ⚠️ 推送返回: {r}")
    except Exception as e:
        print(f"  ❌ 推送失败: {e}")


# ============================================================
# 主流程
# ============================================================
def main():
    start = datetime.now()
    start_time, end_time = get_time_range()
    
    print(f"\n{'='*70}")
    print(f"🚀 春招岗位采集启动 {start.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📅 时间范围: {start_time.strftime('%Y-%m-%d')} ~ {end_time.strftime('%Y-%m-%d')}")
    print(f"{'='*70}")
    print(f"📋 求职条件:")
    print(f"   - 学历: {MAX_EDUCATION}及以下")
    print(f"   - 性别: {PREFERRED_GENDER}")
    print(f"   - 年龄: {MIN_AGE}岁及以上")
    print(f"   - 专业: {'/'.join(PREFERRED_MAJORS)}")
    print(f"   - 地点: 太原/吕梁/山西优先")
    print(f"{'='*70}\n")

    # Step 1: 采集文章列表
    all_articles = []
    for account in ACCOUNTS:
        print(f"📡 采集: {account}")
        arts = fetch_articles_sogou(account, start_time, end_time)
        print(f"   找到 {len(arts)} 篇文章")
        all_articles.extend(arts)

    if not all_articles:
        print("\n⚠️ 未采集到任何文章，停止运行")
        return

    # Step 2: 标题过滤
    filtered = filter_by_title(all_articles)
    print(f"\n🔍 标题过滤: {len(all_articles)} → {len(filtered)} 篇招聘文章\n")

    if not filtered:
        print("⚠️ 过滤后无文章，停止运行")
        return

    # Step 3-6: 正文 + 二维码 + 投递链接 + AI分析 + 二次筛选
    all_jobs = []
    for i, art in enumerate(filtered, 1):
        print(f"[{i}/{len(filtered)}] 处理: {art['title'][:50]}...")
        content, imgs = fetch_article_detail(art["url"])
        
        if not content:
            print(f"   ⚠️ 无有效内容，跳过")
            continue
        
        art["content"] = content
        
        # 第一步：识别二维码
        qr_links = decode_qr_from_images(imgs)
        
        # 第二步：从正文提取投递链接
        text_links = extract_apply_link_from_text(content)
        
        # 合并所有投递链接（去重）
        all_apply_links = list(set(qr_links + text_links))
        
        # 第三步：获取投递页面内容
        apply_page_contents = []
        for link in all_apply_links[:2]:
            if link.startswith("http"):
                page_content = fetch_apply_page_content(link)
                if page_content:
                    apply_page_contents.append(page_content)
            time.sleep(1)
        
        # 第四步：AI分析提取岗位
        print(f"   🤖 AI分析中...")
        jobs = analyze_with_kimi(art, all_apply_links, apply_page_contents)
        
        # 第五步：二次筛选（根据个人条件）
        if jobs:
            print(f"   🔍 执行个人条件筛选...")
            jobs = secondary_filter_jobs(jobs)
        
        all_jobs.extend(jobs)
        time.sleep(2)

    # Step 7: 统计
    valid_jobs    = [j for j in all_jobs if not j.get("excluded") and j.get("score", 0) >= SCORE_THRESHOLD]
    priority_jobs = [j for j in valid_jobs if j.get("score", 0) >= 80]
    
    print(f"\n📊 汇总统计")
    print(f"  - 采集文章: {len(filtered)} 篇")
    print(f"  - 提取岗位: {len(all_jobs)} 个")
    print(f"  - 符合条件: {len(valid_jobs)} 个")
    print(f"  - 优先推荐: {len(priority_jobs)} 个")

    # Step 8: 生成Excel
    excel_path, dl_link = "", ""
    if valid_jobs:
        excel_path = generate_excel(all_jobs)
        print(f"📁 Excel: {excel_path}")
        dl_link = upload_excel(excel_path)
        if dl_link:
            print(f"🔗 下载链接: {dl_link}")

    # Step 9: 推送微信
    elapsed = (datetime.now() - start).seconds
    lines = [
        f"今日采集 **{len(filtered)}** 篇招聘文章，提取 **{len(all_jobs)}** 个岗位，",
        f"符合条件的 **{len(valid_jobs)}** 个，优先推荐 **{len(priority_jobs)}** 个",
        f"\n> 求职条件：{MAX_EDUCATION}及以下学历 | 不限/男性 | {MIN_AGE}岁+ | {'/'.join(PREFERRED_MAJORS)}",
        f"\n> 地点优先：太原 > 吕梁/山西 > 其他地区",
        f"\n⏱ 耗时：{elapsed}秒",
        "",
    ]

    if priority_jobs:
        lines += ["### ✅ 优先推荐岗位（前5条）", ""]
        for job in priority_jobs[:5]:
            sal = f" | {job['salary_info']}" if job.get("salary_info") else ""
            edu = f" | {job['education_req']}" if job.get("education_req") else ""
            gender = f" | {job['gender_req']}" if job.get("gender_req") else ""
            lines.append(
                f"- **{job.get('job_name','')}** · {job.get('employer','')} "
                f"· {job.get('city','')}{sal}{edu}{gender} · {job.get('score',0)}分"
            )
        if len(priority_jobs) > 5:
            lines.append(f"\n...共 {len(priority_jobs)} 个优先岗位")

    if dl_link:
        lines += ["", f"📥 [点击下载Excel岗位清单]({dl_link})（7天有效）"]
    elif valid_jobs:
        lines += ["", "📥 Excel已上传至 GitHub Actions Artifacts，请登录GitHub下载"]

    if not valid_jobs:
        lines += ["", "💬 今日未发现符合条件的岗位，明日继续监控。"]

    push_wechat(
        title=f"🎯 招聘速报 {date.today().strftime('%m/%d')} · {len(valid_jobs)}个岗位 · {len(priority_jobs)}个推荐",
        body="\n".join(lines),
    )
    
    save_state()
    print(f"\n✅ 完成！耗时 {elapsed} 秒\n")


if __name__ == "__main__":
    main()
